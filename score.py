import openpyxl
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor
import subprocess
import os
import numpy as np

current_dir = os.path.dirname(os.path.abspath(__file__))

def open_and_close_excel(file_path):
    try:
        subprocess.Popen([file_path], shell=True)
        subprocess.call("timeout 5", shell=True)
    except Exception as e:
        print(f"Error: {e}")

def fetch_stock_data(stock_code):
    stock = yf.Ticker(stock_code)
    stock_info = stock.info
    
    # Fetch key financial metrics
    trailingEps = stock_info.get('trailingEps', np.nan)
    beta = stock_info.get('beta', np.nan)
    forwardPE = stock_info.get('forwardPE', np.nan)
    returnOnEquity = stock_info.get('returnOnEquity', np.nan)
    returnOnAssets = stock_info.get('returnOnAssets', np.nan)
    grossMargin = stock_info.get('grossMargins', np.nan)
    peRatio = stock_info.get('trailingPE', np.nan)
    pbRatio = stock_info.get('priceToBook', np.nan)
    revenuePerShare = stock_info.get('revenuePerShare', np.nan)
    
    return {
        "stock_code": stock_code,
        "returnOnEquity": returnOnEquity,
        "trailingEps": trailingEps,
        "grossMargin": grossMargin,
        "pbRatio": pbRatio,
        "revenuePerShare": revenuePerShare
    }

def calculate_score(data, weights):
    return (data["returnOnEquity"] * weights["returnOnEquity"] +
            data["trailingEps"] * weights["trailingEps"] +
            data["grossMargin"] * weights["grossMargin"] +
            (1 / data["pbRatio"]) * weights["pbRatio"] +
            data["revenuePerShare"] * weights["revenuePerShare"])

def optimize_weights(stock_data, target_count):
    weights = {
        "returnOnEquity": 0.50,
        "trailingEps": 0.25,
        "grossMargin": 0.15,
        "pbRatio": -0.05,
        "revenuePerShare": 0.05
    }
    
    adjustment_step = 0.01
    max_iterations = 10000
    iteration = 0
    
    while iteration < max_iterations:
        scores = []
        for data in stock_data:
            if (data["returnOnEquity"] > 0 and data["trailingEps"] > 0 and data["grossMargin"] > 0 and data["revenuePerShare"] > 0 and data["pbRatio"] < 0):
                scores.append(calculate_score(data, weights))
            else:
                scores.append(-np.inf)
        
        count_above_zero = sum(1 for score in scores if score > 0)
        
        if count_above_zero == target_count:
            break
        
        if count_above_zero < target_count:
            adjustment_step = abs(adjustment_step)
        else:
            adjustment_step = -abs(adjustment_step)
        
        # Ensure weight constraints are maintained
        if weights["returnOnEquity"] >= weights["trailingEps"] and weights["trailingEps"] > weights["grossMargin"] and weights["grossMargin"] > weights["revenuePerShare"]:
            weights["returnOnEquity"] += adjustment_step
            weights["trailingEps"] += adjustment_step
            weights["grossMargin"] += adjustment_step
            weights["revenuePerShare"] += adjustment_step
            weights["pbRatio"] -= adjustment_step
            
            # Ensure absolute weight of pbRatio is less than ROE and EPS weights
            if abs(weights["pbRatio"]) >= weights["returnOnEquity"]:
                weights["pbRatio"] = -0.5 * weights["returnOnEquity"]
            if abs(weights["pbRatio"]) >= weights["trailingEps"]:
                weights["pbRatio"] = -0.5 * weights["trailingEps"]
        
        total_positive_weight = sum([weights[key] for key in weights if key != "pbRatio"])
        for key in ["returnOnEquity", "trailingEps", "grossMargin", "revenuePerShare"]:
            weights[key] /= total_positive_weight
        
        iteration += 1
    
    return weights

file_path = os.path.join(current_dir, "taiex_mid100_stock_data.xlsx")

with open(os.path.join(current_dir, "tw_stock_codes.txt"), 'r') as file:
    lines = file.read().splitlines()
taiex_mid100_stocks = [line + ".TW" for line in lines]

with ThreadPoolExecutor() as executor:
    stock_data = list(executor.map(fetch_stock_data, taiex_mid100_stocks))

# Optimize weights to find the right combination
target_count = 38
optimal_weights = optimize_weights(stock_data, target_count)

# Print the optimal weights
print("Optimal weights:")
for key, value in optimal_weights.items():
    print(f"{key}: {value:.4f}")

# Calculate scores with optimal weights
results = []
for data in stock_data:
    if (data["returnOnEquity"] > 0 and data["trailingEps"] > 0 and data["grossMargin"] > 0 and data["revenuePerShare"] > 0 and data["pbRatio"] < 0):
        score = calculate_score(data, optimal_weights)
        results.append([data["stock_code"], data["returnOnEquity"], data["trailingEps"], data["grossMargin"], data["pbRatio"], data["revenuePerShare"], score])

# Filter and sort results
filtered_results = [result for result in results if result[-1] > 0]
sorted_results = sorted(filtered_results, key=lambda x: x[-1], reverse=True)
top_38_results = sorted_results[:38]

try:
    workbook = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet_name = "Sheet"
if sheet_name not in workbook.sheetnames:
    sheet = workbook.create_sheet(sheet_name)
else:
    sheet = workbook[sheet_name]

# Define headers
headers = ["股票代碼", "ROE", "EPS", "Gross Margin", "P/B Ratio", "Revenue per Share", "Composite Score"]

for col, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col, value=header)

for i, result in enumerate(top_38_results, start=2):
    for j, value in enumerate(result):
        sheet.cell(row=i, column=j+1, value=value)

workbook.save(file_path)
open_and_close_excel(file_path)

