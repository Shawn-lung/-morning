import openpyxl
import yfinance as yf
from concurrent.futures import ThreadPoolExecutor
import os
import numpy as np

class StockDownloader:
    def __init__(self, stock_list_file, output_file):
        self.stock_list_file = stock_list_file
        self.output_file = output_file
        self.stock_codes = self._load_stock_codes()

    def _load_stock_codes(self):
        with open(self.stock_list_file, 'r') as file:
            lines = file.read().splitlines()
        return [line + ".TW" for line in lines]

    def _fetch_stock_data(self, stock_code):
        stock = yf.Ticker(stock_code)
        stock_info = stock.info
        trailingEps = stock_info.get('trailingEps', np.nan)
        beta = stock_info.get('beta', np.nan)
        forwardPE = stock_info.get('forwardPE', np.nan)
        returnOnEquity = stock_info.get('returnOnEquity', np.nan)
        returnOnAssets = stock_info.get('returnOnAssets', np.nan)
        grossMargin = stock_info.get('grossMargins', np.nan)
        operatingMargin = stock_info.get('operatingMargins', np.nan)
        peRatio = stock_info.get('trailingPE', np.nan)
        pbRatio = stock_info.get('priceToBook', np.nan)
        revenuePerShare = stock_info.get('revenuePerShare', np.nan)
        rtnList = [stock_code, trailingEps, beta, forwardPE, returnOnEquity, returnOnAssets, grossMargin, operatingMargin, peRatio, pbRatio, revenuePerShare]
        print(rtnList)
        return rtnList

    def _clear_sheet(self, sheet):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

    def download_data(self):
        with ThreadPoolExecutor() as executor:
            results = list(executor.map(self._fetch_stock_data, self.stock_codes))
        try:
            workbook = openpyxl.load_workbook(self.output_file)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet_name = "Sheet"
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.create_sheet(sheet_name)
        else:
            sheet = workbook[sheet_name]
            self._clear_sheet(sheet)  

        sheet["A1"] = "股票代碼"
        sheet["B1"] = "EPS"
        sheet["C1"] = "Beta"
        sheet["D1"] = "PE Ratio"
        sheet["E1"] = "ROE"
        sheet["F1"] = "ROA"
        sheet["G1"] = "Gross margin"
        sheet["H1"] = "Operating margin"
        sheet["I1"] = "P/E Ratio"
        sheet["J1"] = "P/B Ratio"
        sheet["K1"] = "Revenue per share"

        for i, result in enumerate(results, start=2):
            sheet[f"A{i}"] = result[0]
            sheet[f"B{i}"] = result[1]
            sheet[f"C{i}"] = result[2]
            sheet[f"D{i}"] = result[3]
            sheet[f"E{i}"] = result[4]
            sheet[f"F{i}"] = result[5]
            sheet[f"G{i}"] = result[6]
            sheet[f"H{i}"] = result[7]
            sheet[f"I{i}"] = result[8]
            sheet[f"J{i}"] = result[9]
            sheet[f"K{i}"] = result[10]

        workbook.save(self.output_file)

    def process(self):
        self.download_data()
