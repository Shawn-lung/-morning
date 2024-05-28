import os
import openpyxl
import pandas as pd
import numpy as np
from sklearn.preprocessing import MinMaxScaler

def calculate_scores(input_file, output_file):
    try:
        workbook = openpyxl.load_workbook(input_file)
    except FileNotFoundError:
        raise FileNotFoundError(f"File {input_file} not found.")

    # 读取工作表
    sheet_name = "Sheet"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name} not found in {input_file}.")

    sheet = workbook[sheet_name]

    # 将Excel工作表转换为DataFrame
    data = sheet.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # 只保留需要的列并去除包含NaN值的行
    required_columns = ["EPS", "Beta", "PE Ratio", "ROE", "Gross margin", "P/B Ratio", "Revenue per share", "Operating margin"]
    df = df.dropna(subset=required_columns)

    # 将数值列转换为数值类型
    for col in required_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # 标准化这些列
    scaler = MinMaxScaler()
    df[['Normalized ROE', 'Normalized EPS', 'Normalized Gross Margin', 'Normalized Revenue per Share', 'Normalized PB Ratio', 'Normalized PE Ratio', 'Normalized Operating Margin']] = scaler.fit_transform(
        df[['ROE', 'EPS', 'Gross margin', 'Revenue per share', 'P/B Ratio', 'PE Ratio', 'Operating margin']]
    )

    # 定义名人选股策略
    def buffet_score(row):
        weights = {
            'Normalized ROE': 0.3,
            'Normalized EPS': 0.2,
            'Normalized Gross Margin': 0.2,
            'Normalized Revenue per Share': 0.2,
            'Normalized PB Ratio': 0.1
        }
        return (row['Normalized ROE'] * weights['Normalized ROE'] +
                row['Normalized EPS'] * weights['Normalized EPS'] +
                row['Normalized Gross Margin'] * weights['Normalized Gross Margin'] +
                row['Normalized Revenue per Share'] * weights['Normalized Revenue per Share'] -
                row['Normalized PB Ratio'] * weights['Normalized PB Ratio'])

    def graham_score(row):
        weights = {
            'Normalized PE Ratio': 0.4,
            'Normalized PB Ratio': 0.4,
            'Normalized EPS': 0.2
        }
        return (row['Normalized PE Ratio'] * weights['Normalized PE Ratio'] +
                row['Normalized PB Ratio'] * weights['Normalized PB Ratio'] +
                row['Normalized EPS'] * weights['Normalized EPS'])

    def o_shaughnessy_score(row):
        weights = {
            'Normalized EPS': 0.4,
            'Normalized PE Ratio': 0.3,
            'Normalized ROE': 0.3
        }
        return (row['Normalized EPS'] * weights['Normalized EPS'] +
                row['Normalized PE Ratio'] * weights['Normalized PE Ratio'] +
                row['Normalized ROE'] * weights['Normalized ROE'])

    def lynch_score(row):
        weights = {
            'Normalized PE Ratio': 0.4,
            'Normalized Revenue per Share': 0.3,
            'Normalized Gross Margin': 0.2,
            'Normalized PB Ratio': 0.1
        }
        return (row['Normalized PE Ratio'] * weights['Normalized PE Ratio'] +
                row['Normalized Revenue per Share'] * weights['Normalized Revenue per Share'] +
                row['Normalized Gross Margin'] * weights['Normalized Gross Margin'] -
                row['Normalized PB Ratio'] * weights['Normalized PB Ratio'])

    def murphy_score(row):
        weights = {
            'Normalized ROE': 0.6,
            'Normalized Operating Margin': 0.4
        }
        return (row['Normalized ROE'] * weights['Normalized ROE'] +
                row['Normalized Operating Margin'] * weights['Normalized Operating Margin'])

    # 计算每只股票的评分
    df['Buffet Score'] = df.apply(buffet_score, axis=1)
    df['Graham Score'] = df.apply(graham_score, axis=1)
    df['O\'Shaughnessy Score'] = df.apply(o_shaughnessy_score, axis=1)
    df['Lynch Score'] = df.apply(lynch_score, axis=1)
    df['Murphy Score'] = df.apply(murphy_score, axis=1)

    # 确保保存所有行
    for i, col in enumerate(df.columns, 1):
        sheet.cell(row=1, column=i).value = col

    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    workbook.save(output_file)

if __name__ == "__main__":
    # 输入文件路径与输出文件路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(current_dir, "taiex_mid100_stock_data.xlsx")
    output_file = os.path.join(current_dir, "taiex_mid100_stock_data_with_scores.xlsx")

    # 计算并保存评分
    calculate_scores(input_file, output_file)

